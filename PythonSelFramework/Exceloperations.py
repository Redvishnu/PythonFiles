import openpyxl
book = openpyxl.load_workbook("C:\\Data\\Book.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=3)
print(cell.value)
sheet.cell(row=2,column=2).value='rahul'
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
Dict={}
for i in range(1, sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row = i,column = j).value)


#logic for to get data for the test case wat we need
for i in range(1,sheet.max_row+1):#tp get rows
    if sheet.cell(row =i ,column =1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            #t0_get columns
            print(sheet.cell(row=i, column=j).value)


#extracting data from excel to dictionary

for i in range(1,sheet.max_row+1):#tp get rows
    if sheet.cell(row =i ,column =1).value == "Testcase2":

        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)
